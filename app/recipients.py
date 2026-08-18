"""Leitura da planilha de funcionarios (.csv / .xlsx).

A planilha de controle da SHILD tem o cabecalho
`EMPRESA | LOGO DA EMPRESA | NOME COMPLETO | RG | CPF | EMAIL | TELEFONE | DATA NASC. | DATA INGRESSO`
mas o mapeamento aceita variacoes: quem digita a planilha nao deveria precisar decorar
o nome exato que o sistema espera. RG, CPF e datas sao ignorados.

O cabecalho nao precisa estar na primeira linha — na planilha da SHILD ele comeca na 5.
Se nao for informado, procuramos sozinhos a linha que parece cabecalho.
"""
from __future__ import annotations

import csv
import io
import re
import unicodedata
from pathlib import Path

from . import fones

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# sinonimos aceitos no cabecalho (sem acento, minusculo)
ALIAS = {
    "email": ("email", "e mail", "e-mail", "mail", "endereco de email", "emails"),
    "empresa": ("empresa", "nome da empresa", "nome empresa", "company", "razao social",
                "razao", "cliente", "organizacao", "unidade"),
    "logo_url": ("logo", "logo url", "logourl", "link da logo", "link logo",
                 "link da logo da empresa", "url da logo", "logo da empresa", "imagem", "link"),
    "nome": ("nome", "nome do funcionario", "funcionario", "colaborador", "nome completo",
             "nome do colaborador"),
    "telefone": ("telefone", "celular", "whatsapp", "whats", "fone", "tel", "numero",
                 "num", "contato", "telefone celular", "telefone whatsapp"),
}


def _norm(h: str) -> str:
    t = unicodedata.normalize("NFKD", str(h or ""))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9 ]+", " ", t.lower()).strip()


def mapear(cabecalho: list[str]) -> dict[str, int]:
    """Cabecalho -> {campo: indice da coluna}. Campo sem coluna simplesmente nao aparece."""
    idx: dict[str, int] = {}
    normalizado = [_norm(h) for h in cabecalho]
    for campo, nomes in ALIAS.items():
        for i, h in enumerate(normalizado):
            if h in nomes and campo not in idx:
                idx[campo] = i
                break
    return idx


def achar_cabecalho(linhas: list[list[str]], dica: int = 0) -> tuple[int, dict[str, int]]:
    """Descobre em qual linha esta o cabecalho — a que reconhecer mais colunas.

    Planilha de RH costuma ter titulo, logo e linhas em branco antes do cabecalho de
    verdade; exigir que ele esteja na primeira linha so gera erro sem motivo.
    """
    if dica and dica <= len(linhas):
        idx = mapear(linhas[dica - 1])
        if idx:
            return dica, idx
    melhor_linha, melhor_idx = 0, {}
    for i, linha in enumerate(linhas[:30], start=1):
        idx = mapear(linha)
        if len(idx) > len(melhor_idx):
            melhor_linha, melhor_idx = i, idx
    return melhor_linha, melhor_idx


def _linhas_csv(caminho: Path) -> list[list[str]]:
    bruto = caminho.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = bruto.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Nao consegui ler o CSV (codificacao desconhecida).")
    amostra = texto[:4096]
    try:
        sep = csv.Sniffer().sniff(amostra, delimiters=",;\t|").delimiter
    except csv.Error:
        sep = ";" if amostra.count(";") > amostra.count(",") else ","
    return [list(r) for r in csv.reader(io.StringIO(texto), delimiter=sep)]


def _linhas_xlsx(caminho: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas = [["" if c is None else str(c).strip() for c in row]
              for row in ws.iter_rows(values_only=True)]
    wb.close()
    return linhas


def ler(caminho: Path, linha_cabecalho: int = 0, ddd_padrao: str = "") -> dict:
    """Devolve {itens, invalidos, duplicados, problemas, colunas, linha_cabecalho}."""
    ext = caminho.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        linhas = _linhas_xlsx(caminho)
    elif ext in (".csv", ".txt"):
        linhas = _linhas_csv(caminho)
    else:
        raise ValueError(f"Formato nao suportado: {ext}. Use .csv ou .xlsx.")

    if not any(any(str(c).strip() for c in ln) for ln in linhas):
        raise ValueError("Arquivo vazio.")

    n_cab, idx = achar_cabecalho(linhas, linha_cabecalho)
    if not idx:
        raise ValueError(
            "Nao reconheci o cabecalho. Esperado algo como: EMPRESA, NOME COMPLETO, "
            "EMAIL, TELEFONE, LOGO DA EMPRESA.")
    if "empresa" not in idx:
        raise ValueError("Nao encontrei a coluna EMPRESA — ela e obrigatoria.")
    if "email" not in idx and "telefone" not in idx:
        raise ValueError("A planilha precisa ter ao menos uma coluna de EMAIL ou TELEFONE.")

    itens, vistos = [], set()
    invalidos = duplicados = 0
    problemas: list[str] = []
    for numero, linha in enumerate(linhas[n_cab:], start=n_cab + 1):
        def g(campo: str) -> str:
            i = idx.get(campo)
            return str(linha[i]).strip() if (i is not None and i < len(linha)
                                             and linha[i] is not None) else ""

        empresa, nome = g("empresa"), g("nome")
        email = g("email").lower()
        if not any((empresa, nome, email, g("telefone"))):
            continue                                    # linha em branco no meio

        telefone, erro_fone = fones.normalizar(g("telefone"), ddd_padrao)
        email_ok = bool(EMAIL_RE.match(email))
        if not email_ok and email:
            problemas.append(f"linha {numero} ({nome or email}): email invalido")
        if not telefone and g("telefone"):
            problemas.append(f"linha {numero} ({nome or 'sem nome'}): {erro_fone}")

        if not email_ok and not telefone:
            invalidos += 1
            if not g("telefone") and not email:
                problemas.append(f"linha {numero} ({nome or 'sem nome'}): sem email nem telefone")
            continue

        chave = f"{_norm(empresa)}|{email or telefone}"
        if chave in vistos:
            duplicados += 1
            continue
        vistos.add(chave)
        itens.append({"chave": chave, "linha": numero, "empresa": empresa,
                      "logo_url": g("logo_url"), "nome": nome,
                      "email": email if email_ok else "", "telefone": telefone,
                      "telefone_erro": erro_fone})

    return {"itens": itens, "invalidos": invalidos, "duplicados": duplicados,
            "problemas": problemas[:20], "linha_cabecalho": n_cab,
            "colunas": {k: linhas[n_cab - 1][v] for k, v in idx.items()
                        if v < len(linhas[n_cab - 1])}}
